import openpyxl

# File --> Workbook --> Sheets --> Rows --> Cells

file = "C:\\Users\\Sreenivas Sreeni\\Documents\\EmpBook2.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active   # sheet = workbook["Sheet1"] --> get active sheet from excel

# rows = sheet.max_row #count no of rows in a excel sheet 6
# cols = sheet.max_column #count no of columns in a excel sheet 4

#Reading all the rows & columns from excel sheet
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value = "welcome"
workbook.save(file)

# multiple files
file = "C:\\Users\\Sreenivas Sreeni\\Documents\\EmpBook2.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # sheet = workbook["Sheet1"] --> get active sheet from excel

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "smith"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "john"
sheet.cell(2,3).value = "manager"

sheet.cell(3,1).value = 789
sheet.cell(3,2).value = "david"
sheet.cell(3,3).value = "developer"

workbook.save(file)  #save the file after enter the data